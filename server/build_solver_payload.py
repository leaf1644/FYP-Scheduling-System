import argparse
import json
import re
from collections import defaultdict
from datetime import datetime, timedelta
from functools import cmp_to_key

from openpyxl import load_workbook


MONTH_MAP = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def parse_args():
    parser = argparse.ArgumentParser(description="Build solver payload from source Excel files.")
    parser.add_argument("students", help="Path to student Excel file")
    parser.add_argument("rooms", help="Path to room Excel file")
    parser.add_argument("availability", help="Path to professor availability Excel file")
    parser.add_argument("--resolve-strategy", choices=["containment", "overlap"], default="overlap")
    parser.add_argument("--timeout-ms", type=int, default=20000)
    parser.add_argument("--output", help="Optional output path")
    return parser.parse_args()


def normalize_header(value):
    return str(value or "").strip().lower().replace("_", "").replace(" ", "")


def pick_value(row, aliases):
    for alias in aliases:
        value = row.get(alias)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def split_list(raw):
    if not raw:
        return []
    return [part.strip() for part in re.split(r"[,;|]", str(raw)) if part and str(part).strip()]


def normalize_professor_id(value):
    text = str(value or "").strip()
    if not text:
        return ""

    compact = re.sub(r"\s+", "", text).upper()
    prefixed = re.match(r"^([A-Z]+)0*(\d+)$", compact)
    if prefixed:
        return f"{prefixed.group(1)}{int(prefixed.group(2)):02d}"

    numeric = re.match(r"^0*(\d+)$", compact)
    if numeric:
        return f"P{int(numeric.group(1)):02d}"

    return compact


def read_rows(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []

    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    output = []
    for row in rows[1:]:
        mapped = {}
        non_empty = False
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else ""
            if value is None:
                value = ""
            if value != "":
                non_empty = True
            mapped[header] = value
        if non_empty:
            output.append(mapped)
    return output


def excel_serial_to_date_key(value):
    base = datetime(1899, 12, 30)
    return (base + timedelta(days=int(float(value)))).strftime("%Y-%m-%d")


def to_date_label(value):
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, (int, float)):
        return excel_serial_to_date_key(value)

    text = str(value or "").strip()
    if not text:
        return ""

    iso_match = re.search(r"(\d{4}-\d{2}-\d{2})", text)
    if iso_match:
        return iso_match.group(1)

    try:
        parsed = datetime.fromisoformat(text)
        return parsed.strftime("%Y-%m-%d")
    except ValueError:
        return text


def get_month_day_key(raw):
    cleaned = re.sub(r"\([^)]*\)", " ", str(raw or "").lower())
    cleaned = cleaned.replace(",", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    if not cleaned:
        return None
    month_token = next((token for token in cleaned.split(" ") if token in MONTH_MAP), None)
    day_match = re.search(r"\b(\d{1,2})\b", cleaned)
    if not month_token or not day_match:
        return None
    return f"{MONTH_MAP[month_token]:02d}-{int(day_match.group(1)):02d}"


def get_day_key(raw):
    iso_match = re.search(r"(\d{4})-(\d{2})-(\d{2})", str(raw or ""))
    if iso_match:
        return f"{iso_match.group(2)}-{iso_match.group(3)}"
    return get_month_day_key(raw) or str(raw or "").strip().lower()


def parse_clock_to_minutes(raw, fallback_meridiem=None):
    cleaned = re.sub(r"\s+", "", str(raw or "").lower())
    match = re.match(r"^(\d{1,2})(?::(\d{2}))?(am|pm)?$", cleaned)
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2) or "0")
    meridiem = match.group(3) or fallback_meridiem

    if not meridiem:
        if hour > 23 or minute > 59:
            return None
        return hour * 60 + minute

    if hour < 1 or hour > 12 or minute > 59:
        return None
    if hour == 12:
        hour = 0
    if meridiem.lower() == "pm":
        hour += 12
    return hour * 60 + minute


def parse_time_range(label):
    normalized = re.sub(r"[–—]", "-", str(label or "").strip())
    match = re.match(
        r"^(.*?)(\d{1,2}(?::\d{2})?\s*(?:am|pm)?)\s*-\s*(\d{1,2}(?::\d{2})?\s*(?:am|pm)?)\s*$",
        normalized,
        re.IGNORECASE,
    )
    if not match:
        return None

    day_prefix = match.group(1).strip()
    day_key = get_day_key(day_prefix)
    start_raw = match.group(2).strip()
    end_raw = match.group(3).strip()
    explicit_start = re.search(r"(am|pm)\s*$", start_raw, re.IGNORECASE)
    explicit_end = re.search(r"(am|pm)\s*$", end_raw, re.IGNORECASE)
    start_minutes = parse_clock_to_minutes(
        start_raw,
        (explicit_start.group(1).lower() if explicit_start else None) or (explicit_end.group(1).lower() if explicit_end else None),
    )
    end_minutes = parse_clock_to_minutes(
        end_raw,
        (explicit_end.group(1).lower() if explicit_end else None) or (explicit_start.group(1).lower() if explicit_start else None),
    )

    if (
        start_minutes is not None
        and end_minutes is not None
        and start_minutes >= end_minutes
        and not explicit_start
        and explicit_end
    ):
        opposite = "pm" if explicit_end.group(1).lower() == "am" else "am"
        alt_start = parse_clock_to_minutes(start_raw, opposite)
        if alt_start is not None and alt_start < end_minutes:
            start_minutes = alt_start

    if start_minutes is None or end_minutes is None or start_minutes >= end_minutes:
        return None

    return {
        "dayKey": day_key,
        "startMinutes": start_minutes,
        "endMinutes": end_minutes,
    }


def compare_slot_label(left, right):
    left_range = parse_time_range(left)
    right_range = parse_time_range(right)
    if left_range and right_range:
        if left_range["dayKey"] != right_range["dayKey"]:
            return -1 if left_range["dayKey"] < right_range["dayKey"] else 1
        if left_range["startMinutes"] != right_range["startMinutes"]:
            return left_range["startMinutes"] - right_range["startMinutes"]
        if left_range["endMinutes"] != right_range["endMinutes"]:
            return left_range["endMinutes"] - right_range["endMinutes"]
    return -1 if left < right else (1 if left > right else 0)


def build_auto_slots(room_rows):
    labels = []
    seen = set()
    for row in room_rows:
        date_label = to_date_label(row.get("Date", row.get("date", "")))
        time_label = pick_value(row, ["Time Slot", "timeSlot", "Time", "time"])
        full_label = " ".join(part for part in [date_label, time_label] if part).strip()
        if full_label and full_label not in seen:
            seen.add(full_label)
            labels.append(full_label)
    labels.sort(key=cmp_to_key(compare_slot_label))
    return [
        {"id": f"AUTO_SLOT_{index + 1:03d}", "timeLabel": label}
        for index, label in enumerate(labels)
    ]


def build_slot_resolvers(slots):
    slot_key_to_id = {}
    slot_time_meta = []
    for slot in slots:
        for key in (slot["id"], slot["timeLabel"]):
            slot_key_to_id[str(key).strip().lower()] = slot["id"]
        parsed = parse_time_range(slot["timeLabel"]) or parse_time_range(slot["id"])
        if parsed:
            slot_time_meta.append({"id": slot["id"], "range": parsed})
    return slot_key_to_id, slot_time_meta


def resolve_token(token, slot_key_to_id, slot_time_meta, resolve_strategy):
    normalized = str(token or "").strip().lower()
    if normalized in slot_key_to_id:
        return [slot_key_to_id[normalized]]

    token_range = parse_time_range(token)
    if not token_range:
        return [str(token)]

    matched = []
    for item in slot_time_meta:
        slot_range = item["range"]
        if token_range["dayKey"] and slot_range["dayKey"] and token_range["dayKey"] != slot_range["dayKey"]:
            continue
        if resolve_strategy == "overlap":
            ok = slot_range["startMinutes"] < token_range["endMinutes"] and slot_range["endMinutes"] > token_range["startMinutes"]
        else:
            ok = slot_range["startMinutes"] >= token_range["startMinutes"] and slot_range["endMinutes"] <= token_range["endMinutes"]
        if ok:
            matched.append(item["id"])
    return matched or [str(token)]


def is_available_cell(value):
    normalized = str(value or "").strip().lower()
    if not normalized:
        return False
    false_values = {"0", "n", "no", "false", "f", "x", "-", "na", "n/a", "none", "nil", "unavailable"}
    true_values = {"1", "y", "yes", "true", "t", "v", "available", "ok", "a"}
    if normalized in false_values:
        return False
    if normalized in true_values:
        return True
    return True


def parse_students(path):
    rows = read_rows(path)
    students = []
    for index, row in enumerate(rows, start=1):
        name = pick_value(row, ["name", "Name", "student", "Student", "students", "Students"])
        if not name:
            continue
        students.append(
            {
                "id": pick_value(row, ["id", "ID", "studentId", "StudentID"]) or f"S{index}",
                "name": name,
                "supervisorId": normalize_professor_id(pick_value(row, ["supervisorId", "SupervisorId", "supervisor", "Supervisor"])),
                "observerId": normalize_professor_id(pick_value(row, ["observerId", "ObserverId", "observer", "Observer"])),
            }
        )
    return students


def parse_rooms(path, slots):
    rows = read_rows(path)
    slot_key_to_id, slot_time_meta = build_slot_resolvers(slots)
    rooms = {}

    for row in rows:
        room_name = pick_value(row, ["venue", "Venue", "room", "Room", "name", "Name"]).strip()
        date_label = to_date_label(row.get("Date", row.get("date", "")))
        time_label = pick_value(row, ["Time Slot", "timeSlot", "Time", "time"])
        slot_label = " ".join(part for part in [date_label, time_label] if part).strip()

        has_existing_assignment = any(
            pick_value(row, aliases)
            for aliases in (["student", "Student"], ["supervisor", "Supervisor"], ["observer", "Observer"])
        )
        if not room_name or not slot_label or has_existing_assignment:
            continue

        resolved = resolve_token(slot_label, slot_key_to_id, slot_time_meta, "containment")
        room = rooms.setdefault(room_name, {"id": room_name, "name": room_name, "capacity": 1, "availableSlotIds": []})
        for slot_id in resolved:
            if slot_id not in room["availableSlotIds"]:
                room["availableSlotIds"].append(slot_id)

    return [room for room in rooms.values() if room["availableSlotIds"]]


def parse_availability(path, slots, resolve_strategy):
    rows = read_rows(path)
    if not rows:
        return {}

    slot_key_to_id, slot_time_meta = build_slot_resolvers(slots)
    headers = list(rows[0].keys())
    fixed_columns = {"id", "professorid", "name", "professorname", "remarks", "remark", "note", "notes"}
    time_columns = [
        header
        for header in headers
        if normalize_header(header) not in fixed_columns and (parse_time_range(header) is not None or str(header).strip().lower() in slot_key_to_id)
    ]

    availability = defaultdict(set)
    for row in rows:
        prof_id = normalize_professor_id(pick_value(row, ["professorId", "ProfessorId", "id", "ID"]))
        if not prof_id:
            continue
        for column in time_columns:
            cell_value = row.get(column)
            if not is_available_cell(cell_value):
                continue
            for slot_id in resolve_token(column, slot_key_to_id, slot_time_meta, resolve_strategy):
                availability[prof_id].add(slot_id)
    return {prof_id: sorted(list(slot_ids)) for prof_id, slot_ids in availability.items()}


def build_room_slots(rooms, slots):
    slot_map = {slot["id"]: slot for slot in slots}
    room_slots = []
    for room in rooms:
        for slot_id in room["availableSlotIds"]:
            slot = slot_map.get(slot_id)
            if not slot:
                continue
            room_slots.append(
                {
                    "id": f"{room['id']}::{slot['id']}",
                    "roomId": room["id"],
                    "roomName": room["name"],
                    "slotId": slot["id"],
                    "timeLabel": slot["timeLabel"],
                }
            )
    return room_slots


def main():
    args = parse_args()
    room_rows = read_rows(args.rooms)
    slots = build_auto_slots(room_rows)
    students = parse_students(args.students)
    rooms = parse_rooms(args.rooms, slots)
    availability = parse_availability(args.availability, slots, args.resolve_strategy)
    room_slots = build_room_slots(rooms, slots)

    payload = {
        "students": students,
        "allRoomSlots": room_slots,
        "profAvailability": availability,
        "profPreferences": {},
        "timeoutMs": args.timeout_ms,
        "meta": {
            "resolveStrategy": args.resolve_strategy,
            "studentCount": len(students),
            "slotCount": len(slots),
            "roomCount": len(rooms),
            "roomSlotCount": len(room_slots),
            "professorCount": len(availability),
        },
    }

    output = json.dumps(payload, ensure_ascii=False)
    if args.output:
        with open(args.output, "w", encoding="utf-8") as handle:
            handle.write(output)
    else:
        print(output)


if __name__ == "__main__":
    main()