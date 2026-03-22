import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


MONTH_MAP = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def parse_args():
    parser = argparse.ArgumentParser(description="Validate an external schedule candidate against source Excel files.")
    parser.add_argument("candidate", help="Path to candidate JSON file")
    parser.add_argument("students", help="Path to student Excel file")
    parser.add_argument("rooms", help="Path to room Excel file")
    parser.add_argument("availability", help="Path to professor availability Excel file")
    parser.add_argument(
        "--resolve-strategy",
        choices=["containment", "overlap"],
        default="overlap",
        help="Professor availability interpretation strategy",
    )
    return parser.parse_args()


def normalize_professor_id(value):
    text = str(value or "").strip()
    if not text:
        return ""
    compact = re.sub(r"\s+", "", text).upper()
    prefixed = re.match(r"^([A-Z]+)0*(\d+)$", compact)
    if prefixed:
        return f"{prefixed.group(1)}{int(prefixed.group(2)):02d}"
    numeric = re.match(r"^0*(\d+)$", compact)
    if numeric:
        return f"P{int(numeric.group(1)):02d}"
    return compact


def pick_value(row, aliases):
    for alias in aliases:
        value = row.get(alias)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def is_available_cell(value):
    normalized = str(value or "").strip().lower()
    if not normalized:
        return False
    false_values = {"0", "n", "no", "false", "f", "x", "-", "na", "n/a", "none", "nil", "unavailable"}
    true_values = {"1", "y", "yes", "true", "t", "v", "available", "ok", "a"}
    if normalized in false_values:
        return False
    if normalized in true_values:
        return True
    return True


def read_rows(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []

    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    output = []
    for row in rows[1:]:
        mapped = {}
        non_empty = False
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else ""
            text = "" if value is None else str(value)
            if text != "":
                non_empty = True
            mapped[header] = text
        if non_empty:
            output.append(mapped)
    return output


def excel_serial_to_date_key(value):
    base = datetime(1899, 12, 30)
    return (base + timedelta(days=int(value))).strftime("%Y-%m-%d")


def get_month_day_key(raw):
    cleaned = re.sub(r"\([^)]*\)", " ", str(raw or "").lower())
    cleaned = cleaned.replace(",", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    if not cleaned:
        return None
    month_token = next((token for token in cleaned.split(" ") if token in MONTH_MAP), None)
    day_match = re.search(r"\b(\d{1,2})\b", cleaned)
    if not month_token or not day_match:
        return None
    return f"{MONTH_MAP[month_token]:02d}-{int(day_match.group(1)):02d}"


def get_day_key(raw):
        iso_match = re.search(r"(\d{4})-(\d{2})-(\d{2})", str(raw or ""))
        if iso_match:
            return f"{iso_match.group(2)}-{iso_match.group(3)}"
        return get_month_day_key(raw)


def parse_clock_to_minutes(raw, fallback_meridiem=None):
    cleaned = re.sub(r"\s+", "", str(raw or "").lower())
    match = re.match(r"^(\d{1,2})(?::(\d{2}))?(am|pm)?$", cleaned)
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2) or "0")
    meridiem = match.group(3) or fallback_meridiem

    if not meridiem:
        if hour > 23 or minute > 59:
            return None
        return hour * 60 + minute

    if hour < 1 or hour > 12 or minute > 59:
        return None
    if hour == 12:
        hour = 0
    if meridiem.lower() == "pm":
        hour += 12
    return hour * 60 + minute


def parse_time_range(label):
    normalized = re.sub(r"[–—]", "-", str(label or "").strip())
    match = re.match(
        r"^(.*?)(\d{1,2}(?::\d{2})?\s*(?:am|pm)?)\s*-\s*(\d{1,2}(?::\d{2})?\s*(?:am|pm)?)\s*$",
        normalized,
        re.IGNORECASE,
    )
    if not match:
        return None

    day_prefix = match.group(1).strip()
    day_key = get_day_key(day_prefix)
    start_raw = match.group(2).strip()
    end_raw = match.group(3).strip()
    explicit_start = re.search(r"(am|pm)\s*$", start_raw, re.IGNORECASE)
    explicit_end = re.search(r"(am|pm)\s*$", end_raw, re.IGNORECASE)
    start_minutes = parse_clock_to_minutes(start_raw, (explicit_start.group(1).lower() if explicit_start else None) or (explicit_end.group(1).lower() if explicit_end else None))
    end_minutes = parse_clock_to_minutes(end_raw, (explicit_end.group(1).lower() if explicit_end else None) or (explicit_start.group(1).lower() if explicit_start else None))

    if start_minutes is None or end_minutes is None or start_minutes >= end_minutes:
        return None

    return {
        "dayKey": day_key,
        "startMinutes": start_minutes,
        "endMinutes": end_minutes,
    }


def assignment_matches_prof_block(date_key, time_slot, block, resolve_strategy):
    fine_range = parse_time_range(f"{date_key} {time_slot}")
    block_range = parse_time_range(block)
    if not fine_range or not block_range:
        return False
    if fine_range["dayKey"] != block_range["dayKey"]:
        return False

    if resolve_strategy == "overlap":
        return fine_range["startMinutes"] < block_range["endMinutes"] and fine_range["endMinutes"] > block_range["startMinutes"]

    return fine_range["startMinutes"] >= block_range["startMinutes"] and fine_range["endMinutes"] <= block_range["endMinutes"]


def build_students(student_path):
    rows = read_rows(student_path)
    students = {}
    for index, row in enumerate(rows, start=1):
        student_label = pick_value(row, ["name", "Name", "student", "Student", "students", "Students"]) or str(index)
        try:
            candidate_id = int(student_label)
        except ValueError:
            candidate_id = index

        students[candidate_id] = {
            "rowIndex": index,
            "student": student_label,
            "supervisor": normalize_professor_id(pick_value(row, ["supervisorId", "SupervisorId", "supervisor", "Supervisor"])),
            "observer": normalize_professor_id(pick_value(row, ["observerId", "ObserverId", "observer", "Observer"])),
        }
    return students


def build_room_availability(room_path):
    workbook = load_workbook(room_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    allowed = set()
    for row in list(sheet.iter_rows(values_only=True))[1:]:
        if not row or row[0] is None:
            continue
        if any(str(value).strip() for value in row[4:7] if value is not None):
            continue
        date_value = row[0]
        date_key = date_value.strftime("%Y-%m-%d") if hasattr(date_value, "strftime") else str(date_value)[:10]
        time_slot = str(row[1]).strip() if row[1] is not None else ""
        venue = str(row[2]).strip() if row[2] is not None else ""
        if date_key and time_slot and venue:
            allowed.add((date_key, time_slot, venue))
    workbook.close()
    return allowed


def build_prof_availability(availability_path):
    rows = read_rows(availability_path)
    if not rows:
        return {}

    headers = list(rows[0].keys())
    fixed_columns = {"id", "professorid", "name", "professorname", "remarks", "remark", "note", "notes"}
    time_columns = [
        header
        for header in headers
        if header.lower().replace("_", "").replace(" ", "") not in fixed_columns and parse_time_range(header) is not None
    ]

    availability = defaultdict(list)
    for row in rows:
        prof_id = normalize_professor_id(pick_value(row, ["professorId", "ProfessorId", "id", "ID"]))
        if not prof_id:
            continue
        for column in time_columns:
            if is_available_cell(row.get(column)):
                availability[prof_id].append(column)
    return availability


def validate_candidate(candidate_path, students_path, rooms_path, availability_path, resolve_strategy):
    candidate = json.loads(Path(candidate_path).read_text(encoding="utf-8"))
    assignments = candidate.get("assignments", [])
    unscheduled = candidate.get("unscheduled", [])
    students = build_students(students_path)
    room_availability = build_room_availability(rooms_path)
    prof_availability = build_prof_availability(availability_path)

    student_counts = Counter(item["student"] for item in assignments)
    duplicate_students = [
        {"student": student, "count": count}
        for student, count in sorted(student_counts.items())
        if count > 1
    ]

    assigned_students = set(student_counts)
    unscheduled_students = {item["student"] for item in unscheduled}
    all_students = set(students)

    room_slot_usage = defaultdict(list)
    professor_slot_usage = defaultdict(list)
    invalid_room_slots = []
    invalid_student_metadata = []
    invalid_prof_availability = []

    for item in assignments:
        student_id = item["student"]
        date_key = excel_serial_to_date_key(item["date"])
        time_slot = item["timeSlot"]
        venue = item["venue"]
        room_slot_usage[(date_key, time_slot, venue)].append(student_id)

        if student_id in students:
            expected = students[student_id]
            if item["supervisor"] != expected["supervisor"] or item["observer"] != expected["observer"]:
                invalid_student_metadata.append({
                    "student": student_id,
                    "expectedSupervisor": expected["supervisor"],
                    "expectedObserver": expected["observer"],
                    "actualSupervisor": item["supervisor"],
                    "actualObserver": item["observer"],
                })

        for professor in (item["supervisor"], item["observer"]):
            professor_slot_usage[(professor, date_key, time_slot)].append(student_id)
            blocks = prof_availability.get(professor, [])
            if not any(assignment_matches_prof_block(date_key, time_slot, block, resolve_strategy) for block in blocks):
                invalid_prof_availability.append({
                    "student": student_id,
                    "professor": professor,
                    "date": date_key,
                    "timeSlot": time_slot,
                })

        if (date_key, time_slot, venue) not in room_availability:
            invalid_room_slots.append({
                "student": student_id,
                "date": date_key,
                "timeSlot": time_slot,
                "venue": venue,
            })

    professor_conflicts = [
        {"professor": professor, "date": date_key, "timeSlot": time_slot, "students": student_ids}
        for (professor, date_key, time_slot), student_ids in sorted(professor_slot_usage.items())
        if len(student_ids) > 1
    ]
    room_conflicts = [
        {"date": date_key, "timeSlot": time_slot, "venue": venue, "students": student_ids}
        for (date_key, time_slot, venue), student_ids in sorted(room_slot_usage.items())
        if len(student_ids) > 1
    ]

    missing_students = sorted(all_students - assigned_students - unscheduled_students)
    assigned_and_unscheduled = sorted(assigned_students & unscheduled_students)

    unique_assigned_students = len(assigned_students)
    is_valid = not any([
        duplicate_students,
        professor_conflicts,
        room_conflicts,
        invalid_room_slots,
        invalid_prof_availability,
        invalid_student_metadata,
        missing_students,
        assigned_and_unscheduled,
        candidate.get("scheduled_count") != unique_assigned_students,
    ])

    return {
        "valid": is_valid,
        "resolveStrategy": resolve_strategy,
        "summary": {
            "scheduledCountClaimed": candidate.get("scheduled_count"),
            "assignmentRows": len(assignments),
            "uniqueAssignedStudents": unique_assigned_students,
            "unscheduledCount": len(unscheduled),
            "duplicateStudentCount": len(duplicate_students),
            "professorConflictCount": len(professor_conflicts),
            "roomConflictCount": len(room_conflicts),
            "invalidRoomSlotCount": len(invalid_room_slots),
            "invalidProfessorAvailabilityCount": len(invalid_prof_availability),
            "invalidStudentMetadataCount": len(invalid_student_metadata),
            "missingStudentCount": len(missing_students),
            "assignedAndUnscheduledCount": len(assigned_and_unscheduled),
        },
        "issues": {
            "duplicateStudents": duplicate_students,
            "professorConflicts": professor_conflicts,
            "roomConflicts": room_conflicts,
            "invalidRoomSlots": invalid_room_slots,
            "invalidProfessorAvailability": invalid_prof_availability,
            "invalidStudentMetadata": invalid_student_metadata,
            "missingStudents": missing_students,
            "assignedAndUnscheduled": assigned_and_unscheduled,
        },
    }


def main():
    args = parse_args()
    result = validate_candidate(
        Path(args.candidate),
        Path(args.students),
        Path(args.rooms),
        Path(args.availability),
        args.resolve_strategy,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()